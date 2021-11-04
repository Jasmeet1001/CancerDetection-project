import matplotlib.pyplot as plt
import matplotlib as mat
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

#Sorting the excel files by Begin Location and rewriting the original one
def SortData(path1p, path2p, ascending_p = True):
    print('Sorting data...this might take some time')
    df1 = pd.read_excel(r"{0}".format(path1p))
    df2 = pd.read_excel(r"{0}".format(path2p))

    sorted1 = df1.sort_values(by = 'Begin Location', ascending = ascending_p)
    sorted2 = df2.sort_values(by = 'Begin Location', ascending = ascending_p)

    #dataframe = pd.DataFrame(sorted1)
    sorted1.to_excel(r"{0}".format(path1p), index = False)
    #dataframe = pd.DataFrame(sorted2)
    sorted2.to_excel(r"{0}".format(path2p), index = False)
    print('Done.\n')

#Storing data in 2 dictionaries and returning them
def storeData(sheetName1, sheetName2):

    #creating dictonaries to store the location and their respective chromosomes
    chromlist1 = defaultdict(list)
    chromlist2 = defaultdict(list)

    #To store the column number of location and chromosomes in the excel file
    colName1 = {}
    colName2 = {}
    colNum = 0
    
    #displaying max no. of rows
    print(f"Total number of rows/entries in {display_path1[:-5]}: {sheetName1.max_row}")
    print(f"Total number of rows/entries in {display_path2[:-5]}: {sheetName2.max_row}")
    print("Processing...")
    
    #Finding the chromosome and location columns in first file
    for cols1 in sheetName1.iter_rows(min_row = 1, max_row = 1, min_col = 1, max_col = sheetName1.max_column, values_only = True):
        for i in range(sheetName1.max_column):
            if (cols1[i].lower() == "chromosome"):
                colName1[cols1[i].lower()] = colNum
            elif (cols1[i].lower() == "begin location"):
                colName2[cols1[i].lower()] = colNum
            
            colNum += 1
    
    #Storing the chromosome and location from first excel file in chromlist1
    for row1 in sheetName1.iter_rows(min_row = 2, min_col = colName1['chromosome'] + 1, max_col = colName2['begin location'] + 1, values_only = True):
        if (row1[0] != None and row1[colName2['begin location'] - colName1['chromosome']] != None):
            chromlist1[int(str(row1[colName2['begin location'] - colName1['chromosome']]))].append(row1[0])

    #clearing the the lists and ressetting the colNum to 0 to store values for second file
    colName1.clear()
    colName2.clear()
    colNum = 0

    #Finding the chromosome and begin location columns in second file
    for cols2 in sheetName2.iter_rows(min_row = 1, max_row = 1, min_col = 1, max_col = sheetName2.max_column, values_only = True):
        for i in range(sheetName2.max_column):
            if (cols2[i].lower() == "chromosome"):
                colName1[cols2[i].lower()] = colNum
            elif (cols2[i].lower() == "begin location"):
                colName2[cols2[i].lower()] = colNum
            
            colNum += 1

    #Storing the chromosome and location from second excel file in chromlist2
    for row2 in sheetName2.iter_rows(min_row = 2, min_col = colName1['chromosome'] + 1, max_col = colName2['begin location'] + 1, values_only = True):
        if (row2[0] != None and row2[colName2['begin location'] - colName1['chromosome']] != None):
            chromlist2[int(str(row2[colName2['begin location'] - colName1['chromosome']]))].append(row2[0])

    #closing the excel files
    workbook1.close()
    workbook2.close()

    return chromlist1, chromlist2

#Finding common location and their respective chromosome
def findCommon(chromList1, chromList2, loc = [], chro = {}):
    print("Finding common...")

    #Function to check and return which dataSet has more entries/values
    def greater():
        return len(chromList1) > len(chromList2)

    #accessing keys from the the dictonary and finding common ones
    if (greater()):
        for key1 in chromList1.keys():
            if (key1 in chromList2.keys()):
                chro[key1] = [chromList1[key1], chromList2[key1]]
    else:
        for key1 in chromList2.keys():
            if (key1 in chromList1.keys()):
                chro[key1] = [chromList1[key1], chromList2[key1]]

    return chro

#Plotting the values returned by the findCommon function
def GraphPlot(commonChro_p, file_name1, file_name2):
    #Data:
    chro_label = ['chr1', 'chr2', 'chr3', 'chr4', 'chr5', 'chr6', 'chr7',
            'chr8', 'chr9', 'chr10', 'chr11', 'chr12', 'chr13', 'chr14',
            'chr15', 'chr16', 'chr17', 'chr18', 'chr19', 'chr20', 'chr21',
            'chr22', 'chr23', 'chr24']
    loc = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    loc1 = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]

    #Assigning common locations founds as per their chromosomes
    ind = 0
    for key in commonChro_p:
            for row in commonChro_p[key]:
                for col in range(len(row)):
                    if (ind == 0):
                        index1 = int(row[col].split('r')[1]) - 1
                        loc[index1] = key
                    else:
                        index2 = int(row[col].split('r')[1]) - 1
                        loc1[index2] = key
                ind += 1
            ind = 0

    barWidth = 0.45

    #creating the bar plot
    fig, ax = plt.subplots(figsize = (16,9))
    plt.tight_layout(pad = 5)

    #setting the x axis limits
    ax.set_xlim([50, 31000000])

    #creating an array of the number of labels present
    y = np.arange(len(chro_label))

    #plotting the bar plot
    rect1 = ax.barh(y + barWidth/2, loc, barWidth, color = 'r', edgecolor = 'grey',
            label = file_name1)
    rect2 = ax.barh(y - barWidth/2, loc1, barWidth, color = 'b', edgecolor = 'grey',
            label = file_name2)

    #setting limits for x-axis ticks and their rotation
    plt.xticks(np.arange(min(loc), 31000000, 900_000), rotation = 75)

    #formatting how the values are displayed
    ax.xaxis.set_major_formatter(mat.ticker.StrMethodFormatter('{x: .0f}'))

    #removing the boundary lines in graph
    sPine = ['top', 'right']
    remove_spine = [ax.spines[i].set_visible(False) for i in sPine]

    #setting the tick values for y axis
    ax.set_yticks(y)
    ax.set_yticklabels(chro_label)

    #add padding to the labels
    ax.xaxis.set_tick_params(pad = 8)
    ax.yaxis.set_tick_params(pad = 10)

    #creating grid lines
    plt.grid(linestyle = '--', linewidth = 0.5)

    ax.set_title('Common Chromosomes and their locations')
    ax.set_xlabel('Locations')
    ax.set_ylabel('Chromosomes')

    ax.legend()

    #Adding annotations to the bar plots
    ax.bar_label(rect1, fmt = '%.0f', padding = 3, label_type = 'edge')
    ax.bar_label(rect2, fmt = '%.0f', padding = 3, label_type = 'edge')

    plt.show()

#Displaying the common locations found along with the chromosome found in both files           
def display():
    print('Displaying...')
    if (len(commonChro) != 0):
        for key in commonChro.keys():
            print(f"Common chromosome found in: \n{display_path1[:-5]}: {commonChro[key][0]}\n{display_path2[:-5]}: {commonChro[key][1]}\npresent at location {key}")
    else:
        return ("No common locations found!")


#DRIVER CODE
path1 = input("Enter path for file 1: ")
path2 = input("Enter path for file 2: ")

sorted_ask = input("Is the data/values to compare sorted?(Y/n) ")

if (sorted_ask.lower() != 'y'):
    SortData(path1,path2)

display_path1 = path1.split('\\')[-1]
display_path2 = path2.split('\\')[-1]

#Loading the excel files in read only mode
workbook1 = load_workbook(filename = r"{0}".format(path1), read_only = True)
workbook2 = load_workbook(filename = r"{0}".format(path2), read_only = True)

#Loading the active sheet in excel file
sheet1 = workbook1.active
sheet2 = workbook2.active

chromoList1, chromoList2 = storeData(sheet1, sheet2)
commonChro = findCommon(chromoList1, chromoList2)

display()

GraphPlot(commonChro, display_path1[:-5], display_path2[:-5])